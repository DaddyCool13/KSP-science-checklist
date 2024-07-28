import re
import openpyxl
from openpyxl import load_workbook

# Savegame auslesen und Science-Werte übernehmen
def get_new_savegame_science():
    # Definiere den Pfad zur Datei
    eingabepfad = r'C:\Games\Kerbal Space Program\saves\letzter Neustart\persistent.sfs'

    # Lese den Inhalt der Datei und speichere ihn in einer Variablen
    with open(eingabepfad, 'r', encoding='utf-8') as datei:
        savegame_inhalt = datei.read()

    # Suche nach dem Beginn der Science-Einträge
    suchbegriff = r"\t\tScience\s*\{\s*id\s*=\s*"
    match = re.search(suchbegriff, savegame_inhalt, re.MULTILINE)

    # Finde den Index des ersten Vorkommens des Suchbegriffs
    start_index = match.start()
    # Lösche alles vor dem Suchbegriff
    savegame_inhalt = savegame_inhalt[start_index:]
    print(f"Suchbegriff '{suchbegriff}' gefunden und Inhalt davor gelöscht.")

    # Suche nach dem Ende der Science-Einträge
    pattern = r'\}\s*(?!\s*Science)'
    match = re.search(pattern, savegame_inhalt, re.DOTALL)
    
    # Finde den Index des ersten Vorkommens des }
    end_index = match.start()
    # Lösche alles nach dem }
    savegame_inhalt = savegame_inhalt[:end_index + 1]
    return savegame_inhalt

def read_savegame(savegame_path):
    with open(savegame_path, 'r') as file:
        lines = file.readlines()
    
    science_entries = []
    inside_science = False

    for i in range(len(lines)):
        if 'Science' in lines[i] and 'id =' in lines[i + 1]:
            inside_science = True
        
        if inside_science:
            science_entries.append(lines[i])
        
        if inside_science and '}' in lines[i] and (i + 1 >= len(lines) or 'Science' not in lines[i + 1]):
            inside_science = False
    
    return ''.join(science_entries)

def save_science_entries(science_data, output_path):
    with open(output_path, 'w') as file:
        file.write(science_data)

def excel_to_html(input_file, output_file, science_data):
    # Lade die Excel-Datei
    wb = load_workbook(input_file)
    ws = wb.active

    # Dictionary mit den Planetenfarben
    planet_colors = {
        "Sun": ("#0018F0", "white"),
        "Moho": ("#4DFFD2", "black"),
        "Eve": ("#B3B3B3", "black"),
        "Gilly": ("#E65C00", "black"),
        "Kerbin": ("#404040", "white"),
        "Mun": ("#B3B3B3", "black"),
        "Minmus": ("#AC39AC", "black"),
        "Duna": ("#FFBF80", "black"),
        "Ike": ("#666666", "black"),
        "Dres": ("#B37A3F", "black"),
        "Jool": ("#3CF000", "black"),
        "Laythe": ("#FFDF80", "black"),
        "Vall": ("#B32400", "black"),
        "Tylo": ("#57B2D1", "black"),
        "Bop": ("#E6E6E6", "black"),
        "Pol": ("#002185", "white"),
        "Eeloo": ("#FFC824", "black")
    }

    def get_color_for_percentage(percentage):
        # Farbskala von weiß (#FFFFFF) bis dunkelgrün (#004d00)
        min_color = (255, 255, 0)  # Gelb
        max_color = (0, 77, 0)     # Dunkelgrün

        # Konvertiere den Prozentwert in einen Wert zwischen 0 und 1
        ratio = percentage / 100.0

        # Berechne die RGB-Werte basierend auf dem Verhältnis
        red = int(min_color[0] * (1 - ratio) + max_color[0] * ratio)
        green = int(min_color[1] * (1 - ratio) + max_color[1] * ratio)
        blue = int(min_color[2] * (1 - ratio) + max_color[2] * ratio)

        return f"rgb({red}, {green}, {blue})"

    # Sammle alle zusammengeführten Zellen im Voraus
    merged_cells = {}
    for merged_cell_range in ws.merged_cells.ranges:
        for row in range(merged_cell_range.min_row, merged_cell_range.max_row + 1):
            for col in range(merged_cell_range.min_col, merged_cell_range.max_col + 1):
                merged_cells[(row, col)] = merged_cell_range

    # Erzeuge die HTML-Datei und füge den Grundaufbau hinzu
    html_content = [
        '''<!DOCTYPE html>
        <html>
        <head>
            <style>
                body {
                    font-family: "Aptos", sans-serif;
                    font-size: 12px; /* Verkleinerte Schriftgröße */
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    table-layout: fixed;
                }
                table, th, td {
                    border: 1px solid black;
                }
                th, td {
                    padding: 5px;
                    text-align: left;
                }
                th.center, td.center {
                    text-align: center;
                }
                thead th {
                    position: sticky;
                    top: 0;
                    z-index: 1;
                    background-color: #f1f1f1;
                }
                .rotate {
                    writing-mode: vertical-rl;
                    transform: scale(-1, -1);
                    font-size: 24px;
                    height: 100px; /* Anpassen der Höhe nach Bedarf */
                }
            </style>
        </head>
        <body>
        <table>
        <thead>
        '''
    ]

    # Iteriere durch die Zeilen und Spalten der Excel-Tabelle
    for i, row in enumerate(ws.iter_rows()):
        row_content = ['<tr>']
        for j, cell in enumerate(row):
            cell_value = cell.value if cell.value is not None else ""
            cell_style = ""
            cell_class = ""

            # Überprüfe, ob die Zelle Teil eines Merged-Cell-Bereichs ist
            if (cell.row, cell.column) in merged_cells:
                merged_cell_range = merged_cells[(cell.row, cell.column)]
                if (cell.row, cell.column) == (merged_cell_range.min_row, merged_cell_range.min_col):
                    colspan = merged_cell_range.size['columns']
                    rowspan = merged_cell_range.size['rows']

                    if cell_value == 'x':  # Highlight merged cells with 'x' in grey, but don't display 'x'
                        cell_value = ""
                        cell_style = 'style="background-color: #595959; color: white"'
                    elif isinstance(cell_value, (int, float)) and cell_value == 0:  # Convert 0 to 0%
                        cell_value = "0 %"
                        percentage = float(cell_value.replace(" %", "").strip())
                        cell_style = f'style="background-color: {get_color_for_percentage(percentage)}"'
                    elif j == 0 and cell_value in planet_colors:  # Apply planet colors for the first column based on planet name
                        bg_color, font_color = planet_colors[cell_value]
                        cell_style = f'style="background-color: {bg_color}; color: {font_color}"'
                        cell_class = 'class="rotate"'
                    
                    if j >= 3:
                        cell_class = 'class="center"'

                    row_content.append(f'<td colspan="{colspan}" rowspan="{rowspan}" {cell_class} {cell_style}>{cell_value}</td>')
            else:
                if j == 0 and cell_value in planet_colors:  # Apply planet colors for the first column based on planet name
                    bg_color, font_color = planet_colors[cell_value]
                    cell_style = f'style="background-color: {bg_color}; color: {font_color}"'
                    cell_class = 'class="rotate"'
                elif cell_value == 'x':  # Highlight cells with 'x' in grey, but don't display 'x'
                    cell_value = ""
                    cell_style = 'style="background-color: #595959; color: white"'
                elif isinstance(cell_value, (int, float)) and cell_value == 0:  # Convert 0 to 0%
                    cell_value = "0 %"
                    percentage = float(cell_value.replace(" %", "").strip())
                    cell_style = f'style="background-color: {get_color_for_percentage(percentage)}"'
                
                if j >= 3:
                    cell_class = 'class="center"'

                if i == 0:
                    row_content.append(f'<th {cell_class} {cell_style}>{cell_value}</th>')
                else:
                    row_content.append(f'<td {cell_class} {cell_style}>{cell_value}</td>')

        row_content.append('</tr>')
        html_content.append(''.join(row_content))
        if i == 0:
            html_content.append('</thead><tbody>')

    # Schließe die HTML-Tags
    html_content.append('''
    </tbody>
    </table>
    </body>
    </html>
    ''')

    # Schreibe den HTML-Inhalt in die Datei
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(''.join(html_content))

# Pfade zur Excel-Datei und zur zu erzeugenden HTML-Datei
input_file = r'C:\Users\danie\OneDrive\Software\KSP Science Checklist\full science table.xlsx'
output_file = r'C:\Users\danie\OneDrive\Software\KSP Science Checklist\KSP Science Checklist.html'
savegame_file = r'C:\Games\Kerbal Space Program\saves\letzter Neustart\persistent.sfs'
science_output_file = r'C:\Users\danie\OneDrive\Software\KSP Science Checklist\science_entries.txt'

# Lese und speichere die Science-Einträge aus der Savegame-Datei
science_data = read_savegame(savegame_file)
save_science_entries(science_data, science_output_file)

# Funktion aufrufen, um die HTML-Datei zu erstellen
excel_to_html(input_file, output_file, science_data)
